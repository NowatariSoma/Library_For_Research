# 参考
# https://lemon818.com/pythop-excel/
# Python で Excel ファイルを新規作成し、書き込む方法．

import openpyxl as excel
import os

#カレントディレクトリを現在のディレクトリに移動
#https://note.nkmk.me/python-script-file-path/
os.chdir(os.path.dirname(os.path.abspath(__file__)))
msg_A1 = "Helllo World"
msg_B2 = "あいうえお"

wbname = "test.xlsx"
wb = excel.Workbook()
ws = wb.active
# 書き込み処理
ws["A1"] = msg_A1
ws["B2"] = msg_B2
wb.save(wbname)