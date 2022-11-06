import openpyxl as xl
import os
import glob

os.chdir(os.path.dirname(os.path.abspath(__file__)))
input_P = "excel_input/"
input_PL = glob.glob(input_P + "*.xlsx")

print(input_PL)

wbname = "excel_output/log.xlsx"
wb_out = xl.Workbook()
ws_out = wb_out.active

for file_P in input_PL:
    print(file_P)
    wb_in = xl.load_workbook(filename=file_P)
    ws_in = wb_in.worksheets[0]
    sub_FN = os.path.basename(os.path.splitext(os.path.basename(file_P))[0])
    for i in range(10):
        ws_out.cell(row=i+1,column=int(sub_FN)).value = ws_in.cell(row=i+1,column=1).value
    
wb_out.save(wbname)