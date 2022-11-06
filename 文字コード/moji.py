import openpyxl as xl
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# 文字コードを変換する関数
def change(string):
    tmp = string.encode('shift_jis', "ignore")
    utf = tmp.decode('utf-8', "ignore")
    print(utf)
    return utf

# 変換するファイル
inputfile = 'track.xlsx'
wb_in = xl.load_workbook(filename=inputfile)
ws_in = wb_in.worksheets[0]

i = 1
while True:
    if ws_in.cell(row=1,column=i).value == None:
        break
    else:
        print(ws_in.cell(row=1,column=i).value is None)
        ws_in.cell(row=1,column=i).value = change(ws_in.cell(row=1,column=i).value)
        print(ws_in.cell(row=1,column=i).value)
        i += 1

wb_in.save('t.xlsx')